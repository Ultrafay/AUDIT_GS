import io
from openpyxl import load_workbook
from services.openai_extractor import SalesOrderData, SalesInvoiceData, GDNData

def _aggregate_line_items(line_items, qty_attr: str, fallback_total_qty, fallback_total_amount):
    """Return (total_qty, total_amount, is_single_line). If total_qty/amount not extracted,
    sum them from line items. is_single_line is True only when exactly one line item exists."""
    is_single_line = (len(line_items) == 1) if line_items else False
    
    total_qty = fallback_total_qty
    if total_qty is None and line_items:
        qtys = [getattr(item, qty_attr) for item in line_items if getattr(item, qty_attr, None) is not None]
        if qtys:
            total_qty = sum(qtys)

    total_amount = fallback_total_amount
    if total_amount is None and line_items:
        amounts = [getattr(item, 'amount') for item in line_items if getattr(item, 'amount', None) is not None]
        if amounts:
            total_amount = sum(amounts)
            
    return total_qty, total_amount, is_single_line

def populate_template(
    sales_orders: list[SalesOrderData],
    invoices: list[SalesInvoiceData],
    gdns: list[GDNData],
) -> bytes:
    print("[ExcelExport] === populate_template v2 (multi-line fix) called ===")
    template_path = "templates/revenue_audit_3tab.xlsx"
    wb = load_workbook(template_path)
    
    def set_cell(sheet, row, col, value):
        if value is not None:
            sheet.cell(row=row, column=col).value = value

    # Sales Order
    so_sheet = wb["Sales Order"]
    if len(sales_orders) > 10:
        print(f"[ExcelExport] Truncating SalesOrders to 10 rows (got {len(sales_orders)})")
    
    for i, so in enumerate(sales_orders[:10], start=3):
        total_qty, total_amount, is_single_line = _aggregate_line_items(
            so.line_items, 'quantity', so.total_quantity, getattr(so, 'total_amount', None)
        )
        
        doc_ref = so.so_number or f"Row {i}"
        if not is_single_line:
            num_items = len(so.line_items) if so.line_items else 0
            print(f"[ExcelExport] Sales Order {doc_ref}: multi-line ({num_items} items), writing total_amount={total_amount} directly to col F")
        else:
            print(f"[ExcelExport] Sales Order {doc_ref}: single-line")
            
        rate = so.line_items[0].rate if is_single_line and getattr(so.line_items[0], 'rate', None) is not None else None
        
        set_cell(so_sheet, i, 2, so.so_number)        # B
        set_cell(so_sheet, i, 3, so.customer_name)    # C
        set_cell(so_sheet, i, 4, total_qty)           # D
        set_cell(so_sheet, i, 5, rate)                # E
        set_cell(so_sheet, i, 7, so.notes)            # G
        
        if not is_single_line and total_amount is not None:
            # Override column F (formula =D*E) value with a direct numeric total_amount
            set_cell(so_sheet, i, 6, total_amount)    # F

    # Sales Invoice
    inv_sheet = wb["Sales Invoice"]
    if len(invoices) > 10:
        print(f"[ExcelExport] Truncating SalesInvoices to 10 rows (got {len(invoices)})")
        
    for i, inv in enumerate(invoices[:10], start=3):
        total_qty, total_amount, is_single_line = _aggregate_line_items(
            inv.line_items, 'quantity', inv.total_quantity, getattr(inv, 'total_amount', None)
        )
        
        doc_ref = inv.invoice_number or f"Row {i}"
        if not is_single_line:
            num_items = len(inv.line_items) if inv.line_items else 0
            print(f"[ExcelExport] Sales Invoice {doc_ref}: multi-line ({num_items} items), overriding col G & H with {total_amount}")
        else:
            print(f"[ExcelExport] Sales Invoice {doc_ref}: single-line")
            
        rate = inv.line_items[0].rate if is_single_line and getattr(inv.line_items[0], 'rate', None) is not None else None
        
        set_cell(inv_sheet, i, 2, inv.customer_name)  # B
        set_cell(inv_sheet, i, 3, inv.invoice_number) # C
        set_cell(inv_sheet, i, 4, inv.invoice_date)   # D
        set_cell(inv_sheet, i, 5, total_qty)          # E
        set_cell(inv_sheet, i, 6, rate)               # F
        set_cell(inv_sheet, i, 8, total_amount)       # H
        set_cell(inv_sheet, i, 10, inv.notes)         # J
        
        if not is_single_line and total_amount is not None:
            # Provide total_amount plainly in col G (overriding the formula) and col H to avoid "Difference" formula conflicts
            set_cell(inv_sheet, i, 7, total_amount)   # G

    # GDN
    gdn_sheet = wb["GDN"]
    if len(gdns) > 10:
        print(f"[ExcelExport] Truncating GDNs to 10 rows (got {len(gdns)})")
        
    for i, gdn in enumerate(gdns[:10], start=3):
        total_qty, total_amount, is_single_line = _aggregate_line_items(
            gdn.line_items, 'quantity_delivered', gdn.total_quantity_delivered, getattr(gdn, 'total_amount', None)
        )
        
        doc_ref = gdn.gdn_reference or f"Row {i}"
        if not is_single_line:
            num_items = len(gdn.line_items) if gdn.line_items else 0
            print(f"[ExcelExport] GDN {doc_ref}: multi-line ({num_items} items)")
        else:
            print(f"[ExcelExport] GDN {doc_ref}: single-line")
            
        set_cell(gdn_sheet, i, 2, gdn.customer_name)      # B
        set_cell(gdn_sheet, i, 3, gdn.delivered_date)     # C
        set_cell(gdn_sheet, i, 4, total_qty)              # D
        set_cell(gdn_sheet, i, 5, gdn.gdn_reference)      # E
        set_cell(gdn_sheet, i, 6, gdn.notes)              # F

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
